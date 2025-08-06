

import openpyxl


def get_row_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_row


def get_column_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path,sheet_name,row_number,column_number):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.cell(row_number,column_number).value

def set_cell_data(path,sheet_name,row_number,column_number,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    sheet.cell(row_number,column_number).value=data
    workbook.save(path)

def get_data_from_excel(path,sheet_name):
    final_list=[]
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    total_row=sheet.max_row
    total_col=sheet.max_column

    for r in range(2,total_row+1):
        row_list=[]
        for col in range(1,total_col+1):
            row_list.append(sheet.cell(r,col).value)
        final_list.append(row_list)
    return final_list